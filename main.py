from ctypes import cdll
import requests
import json
import csv
import openpyxl


url = "https://api.paylocity.com/IdentityServer/connect/token"

payload='grant_type=client_credentials&scope=WebLinkAPI'
auth = {
  'Authorization': '[access_token]',
  'Content-Type': 'application/x-www-form-urlencoded'
}

response = requests.request("POST", url, headers=auth, data=payload)

headers = {
    'Authorization': 'Bearer {}'.format(response.json()['access_token'])
}

check_date = "YYYY-MM-DD"
path = "deductions-{}.xlsx".format(check_date)
ded_code = "insert_deduction_code"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

employees_with_deductions = []

for row in range(2, sheet_obj.max_row + 1):
    pcid = sheet_obj.cell(row = row, column = 1)
    tpid = sheet_obj.cell(row = row, column = 2)
    employees_with_deductions.append((pcid.value, tpid.value))

def get_deduction(pcid, tpid):
    # url = "https://api.paylocity.com/api/v1/companies/{}/employees/{}/deductions/{}".format("140706", tpid, ded_code)
    url = "https://api.paylocity.com/api/v2/companies/{companyId}/employees/{employeeId}/paystatement/details/{year}/{checkDate}".format(companyId=pcid, employeeId=tpid, year="2024", checkDate=check_date)

    response = requests.get(url, headers=headers)
    if response.status_code != 204:
        try:
            response_dict = json.loads(response.text)
        except:
            return 0
        index = (len(response_dict) - 1)
        try:
            while index >= 0:
                if response.json()[index]['detCode'] == ded_code:
                    return response.json()[index]['amount']
                else:
                    index -= 1
            return 0
        except:
            return "NotFound"
    else:
        return "err204"
        

write_file = open("Paylocity_actuals_{}.csv".format(check_date), "w")
file_writer = csv.writer(write_file)
file_writer.writerow(["PCID", "TPID", "Deduction Amount"])

for row in employees_with_deductions:
    pcid = row[0]
    tpid = row[1]
    try:
        deduction_amount = get_deduction(pcid, tpid)
    except:
        deduction_amount = "err"
    print("----------")
    print(tpid)
    print(deduction_amount)
    file_writer.writerow([pcid, tpid, deduction_amount])
